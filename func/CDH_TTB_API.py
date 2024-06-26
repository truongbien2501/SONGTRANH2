import pandas as pd
import numpy as np
from datetime import datetime,timedelta
from func.Seach_file import tim_file,read_txt,vitridat
from tkinter import messagebox
from openpyxl import load_workbook
from win32com import client
import time
import io
def TTB_API_mucnuoc():
    now = datetime.now()
    kt = datetime(now.year,now.month,now.day,now.hour)
    bd = kt - timedelta(days=1)
    data = pd.DataFrame()
    data['time'] = pd.date_range(bd,kt,freq='T')
    tram = pd.read_csv('ts_id/TTB_H_ODA.txt')
    for item in zip(tram.Matram,tram.tentram,tram.TAB):
    # print(item[0],item[2],item[1])
        pth = 'http://113.160.225.84:2018/API_TTB/XEM/solieu.php?matram={}&ten_table={}&sophut=1&tinhtong=0&thoigianbd=%27{}%2000:00:00%27&thoigiankt=%27{}%2023:59:00%27'
        pth = pth.format(item[0],item[2],bd.strftime('%Y-%m-%d'),kt.strftime('%Y-%m-%d'))
        df = pd.read_html(pth)
        df[0].rename(columns={"thoi gian":'time','so lieu':item[1]},inplace=True)
        df = df[0].drop('Ma tram',axis=1)
        df['time'] = pd.to_datetime(df['time'])
        data = data.merge(df,how='left',on='time')
    data.set_index('time',inplace=True)
    data = data[data.index.minute == 0]

    # data = data[['Son Giang','Tra Khuc','An Chi','Song Ve','Chau O','Tra Cau','Binh Dong','Dung Quat Idro']]*100
    data =data.astype(float)
    return data

def TTB_API_mucnuoc10day():
    now = datetime.now()
    kt = datetime(now.year,now.month,now.day,now.hour)
    bd = kt - timedelta(days=11)
    data = pd.DataFrame()
    data['time'] = pd.date_range(bd,kt,freq='min')
    tram = pd.read_csv('ts_id/Qsongtranh.txt')
    for item in zip(tram.Matram,tram.tentram,tram.TAB):
    # print(item[0],item[2],item[1])
        pth = 'http://113.160.225.84:2018/API_TTB/XEM/solieu.php?matram={}&ten_table={}&sophut=60&tinhtong=0&thoigianbd=%27{}%2000:00:00%27&thoigiankt=%27{}%2023:59:00%27'
        pth = pth.format(item[0],item[2],bd.strftime('%Y-%m-%d'),kt.strftime('%Y-%m-%d'))
        df = pd.read_html(pth)
        df[0].rename(columns={"thoi gian":'time','so lieu':item[1]},inplace=True)
        df = df[0].drop('Ma tram',axis=1)
        df['time'] = pd.to_datetime(df['time'])
        data = data.merge(df,how='left',on='time')
    data.set_index('time',inplace=True)
    data = data[data.index.minute == 0]

    # data = data[['Son Giang','Tra Khuc','An Chi','Song Ve','Chau O','Tra Cau','Binh Dong','Dung Quat Idro']]*100
    data =data.astype(float)
    # print(data)
    return data
import paramiko
def read_muadb_sever_hochua(tg_db,tenho):
    hostname = '203.209.181.171'
    port = 22
    username = 'mpi'
    password = 'mpi@1234'
    # Tạo kết nối SSH
    client = paramiko.SSHClient()
    client.load_system_host_keys()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(hostname, port, username, password)
    # Đường dẫn tới tệp tin .txt trên máy chủ
    tg = datetime.now()
    # tg = (ngay - timedelta(days=1)).strftime('%d%m%Y')
    obs = ['12', '00']
    ngaylayfile = [tg,tg-timedelta(days=1)]
    thoat = False
    for ngay in ngaylayfile:
        if thoat:
            break
        for zz in obs:        
            try:
                if tenho !='A VƯƠNG':
                    remote_file_path = '/home/disk2/KQ_WRF72h/' + ngay.strftime('%d%m%Y') + '/Hinh_{}z_36h/QuangNam_{}z_3k36h.txt'.format(zz,zz)
                else:
                    remote_file_path = '/home/disk2/KQ_WRF72h/' + ngay.strftime('%d%m%Y') + '/{}z/TTB_{}z_36h.txt'.format(zz,zz)
                    thoigianbatdau = datetime(ngay.year,ngay.month,ngay.day,int(zz))
                    # print(thoigianbatdau)
                # Đọc nội dung của tệp tin từ máy chủ
                sftp = client.open_sftp()
                remote_file = sftp.open(remote_file_path)
                file_contents = remote_file.read().decode()
                thoat= True
                print(remote_file_path)
                break
            except:
                pass
    
    # print(file_contents)
    # Đóng kết nối SSH
    remote_file.close()
    sftp.close()
    client.close()
    df = pd.read_csv(io.StringIO(file_contents), delimiter=r"\s+")
    if tenho !='A VƯƠNG':
        df = df[1:]
        df = df.T
        df.columns = df.iloc[0]
        df = df[1:]
    else:
        # df = df[1:]
        df = df.T
        df.columns = df.iloc[0]
        df = df[1:]
        # print(df)
        # print(df.T)
    if tenho =='SÔNG TRANH 2':
        df =df[['TraBui(ST2)','TraCang(ST2)','TraDon(ST2)','TraGiac(ST2)','TraLeng(ST2)','TraLinh(ST2)','TraMai(ST2)','UBNDNTM(ST2)','Dap(ST2)','TraNam2(ST2)','TraVan(ST2)']]
        # df.columns = ['Trà Bui','Trà Cang','Trà Dơn','Trà Giác','Trà Leng','Trà Linh','Trà Mai','UBNDNTM','Đập chính','Trà Nam','Trà Vân']
    elif tenho =='A VƯƠNG':
        # print(df)
        df = df.reset_index(False)
        # df =df[['TraBui(ST2)','TraCang(ST2)','TraDon(ST2)','TraGiac(ST2)','TraLeng(ST2)','TraLinh(ST2)','TraMai(ST2)','UBNDNTM(ST2)','Dap(ST2)','TraNam2(ST2)','TraVan(ST2)']]
        df = df[['AV1','AV2','AV3','AV4','AV5','AV6','HIEN']]
        # name_viet = ['Đập tràn A Vương','UBND Ab Vương','Đồn biên phòng A Nông','UBND Huyện Tây Giang','UBND Xã Dang','Trạm Xã A Tep','Trạm Xã A Rooi','Trạm UBND Xã Blahee']
        # df.columns = ['Đập tràn A Vương','UBND Ab Vương','Đồn biên phòng A Nông','UBND Huyện Tây Giang','UBND Xã Dang','Trạm Xã A Tep','Hien']
    elif tenho =='SÔNG BUNG 2':
        df =df[['DapSBung2','TrHySBung2','NMSongBung2','GaRiSBung2']]
        # df.columns = ['Đập SB2','TrHy','Chơm','A Xan']
    elif tenho =='SÔNG BUNG 4':
        df =df[['DonBQNGiang','ChaVaNMDHSB4','DapSBung4','ZuoihSBung4','TrHySBung2','LaDeeSBung4','CuakhauNG']]
        # df.columns = ['ĐăkPring','Chalval','Đầu mối','Zuôi','TrHy','LaDee','Đak Ốc']
    
    if tenho !='A VƯƠNG':
        df.dropna(inplace=True)
        df = df.reset_index(drop=True)
        first_line = file_contents[:file_contents.index('Time')-1]
        pp = first_line.split(" ")
        ngaythangnam = pp[-1].replace('\n','')
        gio = pp[-3].replace('h','')
        thoidoanmua = pp[8].replace('h','')
        tgbd  = datetime.strptime(ngaythangnam + " " + gio, "%d-%m-%Y %H") + timedelta(hours=int(thoidoanmua))
        df.insert(0,'time',pd.date_range(tgbd,periods=(df.shape[0]),freq=pp[8]))
    else:
        # print(thoigianbatdau)
        df.insert(0,'time',pd.date_range(thoigianbatdau+ timedelta(hours=10),periods=(df.shape[0]),freq='3h')) 
    # print(df)
    df = df[(df['time']>= datetime.now()) & (df['time']<= datetime.now() + timedelta(hours=tg_db))]
    df.iloc[:,1:] = df.iloc[:,1:].astype(float)
    print(df)
    df =df.iloc[:,1:].sum()
    df = df.to_frame().T
    return df

def TTB_API_mucnuoc():
    now = datetime.now()
    kt = datetime(now.year,now.month,now.day,now.hour)
    bd = kt - timedelta(days=1)
    data = pd.DataFrame()
    data['time'] = pd.date_range(bd,kt,freq='min')
    tram = pd.read_csv('ts_id/TTB_H_ODA.txt')
    for item in zip(tram.Matram,tram.tentram,tram.TAB):
    # print(item[0],item[2],item[1])
        pth = 'http://113.160.225.84:2018/API_TTB/XEM/solieu.php?matram={}&ten_table={}&sophut=1&tinhtong=0&thoigianbd=%27{}%2000:00:00%27&thoigiankt=%27{}%2023:59:00%27'
        pth = pth.format(item[0],item[2],bd.strftime('%Y-%m-%d'),kt.strftime('%Y-%m-%d'))
        df = pd.read_html(pth)
        df[0].rename(columns={"thoi gian":'time','so lieu':item[1]},inplace=True)
        df = df[0].drop('Ma tram',axis=1)
        df['time'] = pd.to_datetime(df['time'])
        data = data.merge(df,how='left',on='time')
    data.set_index('time',inplace=True)
    data = data[data.index.minute == 0]

    # data = data[['Son Giang','Tra Khuc','An Chi','Song Ve','Chau O','Tra Cau','Binh Dong','Dung Quat Idro']]*100
    data =data.astype(float)
    return data
def TTB_API_mucnuoc_lu():
    now = datetime.now()
    kt = datetime(now.year,now.month,now.day,now.hour)
    bd = kt - timedelta(days=1)
    data = pd.DataFrame()
    data['time'] = pd.date_range(bd,kt,freq='min')
    tram = pd.read_csv('ts_id/Hlu.txt')
    for item in zip(tram.Matram,tram.tentram,tram.TAB):
    # print(item[0],item[2],item[1])
        pth = 'http://113.160.225.84:2018/API_TTB/XEM/solieu.php?matram={}&ten_table={}&sophut=1&tinhtong=0&thoigianbd=%27{}%2000:00:00%27&thoigiankt=%27{}%2023:59:00%27'
        pth = pth.format(item[0],item[2],bd.strftime('%Y-%m-%d'),kt.strftime('%Y-%m-%d'))
        df = pd.read_html(pth)
        df[0].rename(columns={"thoi gian":'time','so lieu':item[1]},inplace=True)
        df = df[0].drop('Ma tram',axis=1)
        df['time'] = pd.to_datetime(df['time'])
        data = data.merge(df,how='left',on='time')
    data.set_index('time',inplace=True)
    data = data[data.index.minute == 0]

    # data = data[['Son Giang','Tra Khuc','An Chi','Song Ve','Chau O','Tra Cau','Binh Dong','Dung Quat Idro']]*100
    data =data.astype(float)
    return data

def TTB_API_mua():
    now = datetime.now()
    kt = datetime(now.year,now.month,now.day,7)
    bd = kt - timedelta(days=1)
    data = pd.DataFrame()
    data['time'] = pd.date_range(bd,kt,freq='T')
    tram = pd.read_csv('TS_ID/TTB/TTB_MUA_ODA.txt')
    tram['TAB1'] = tram['TAB'].replace(['mua_oday_thuyvan', 'mua_oday_khituong','mua_oday_domua'], 'ODA')
    order = ['ODA', 'hanquoc_mua', 'vrain_mua', 'mua_wb5']
    tram['TAB_category'] = pd.Categorical(tram['TAB1'], categories=order, ordered=True)
    tram = tram.sort_values(by=['TAB_category', 'tentram'])
    tram = tram.drop(columns=['TAB_category','TAB1'])
    # print(tram)
    for item in zip(tram.Matram,tram.tentram,tram.TAB):
    # print(item[0],item[2],item[1])
        pth = 'http://113.160.225.84:2018/API_TTB/XEM/solieu.php?matram={}&ten_table={}&sophut=1&tinhtong=0&thoigianbd=%27{}%2000:00:00%27&thoigiankt=%27{}%2023:59:00%27'
        pth = pth.format(item[0],item[2],bd.strftime('%Y-%m-%d'),kt.strftime('%Y-%m-%d'))
        df = pd.read_html(pth)
        df[0].rename(columns={"thoi gian":'time','so lieu':item[1]},inplace=True)
        df = df[0].drop('Ma tram',axis=1)
        df['time'] = pd.to_datetime(df['time'])
        data = data.merge(df,how='left',on='time')
    data.set_index('time',inplace=True)
    muagio = data.rolling(60,min_periods=1).sum()
    muagio = muagio[muagio.index.minute == 0]
    epsilon = 1e-10
    muagio = muagio.applymap(lambda x: 0 if abs(x) < epsilon else x)    
    muagio =muagio.astype(float)
    return muagio

def CDH_API_mucnuoc():
    now = datetime.now()
    kt = datetime(now.year,now.month,now.day,7)
    bd = kt - timedelta(days=1)
    dfname = pd.read_csv('TS_ID/CDH/CDH_H_QNGA.txt',sep="\s+",header=None)
    data = pd.DataFrame()
    data['time'] = pd.date_range(bd,kt,freq='H')
    
    for tsid in zip(dfname[0],dfname[1]):
        url = 'https://cdh.vnmha.gov.vn/KiWIS/KiWIS?http://slportal.kttv.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request=getTimeseriesValues&datasource=0&format=html&ts_id={}&from={}&to={}'
        df = pd.read_html(url.format(tsid[0],bd.strftime('%Y-%m-%d'),kt.strftime('%Y-%m-%d')))
        for i in df:
            i = i.iloc[4:,:].dropna()
            i.rename(columns={0:'time',1:tsid[1]},inplace=True)
            i['time'] = pd.to_datetime(i['time'])
            i['time'] = i['time'].dt.strftime('%Y-%m-%d %H:%M:%S')
            i['time'] = pd.to_datetime(i['time'])
            data = data.merge(i,how='left',on='time')
    data.set_index('time',inplace=True)
    data =data.astype(float)
    data['sg_oda'].update(data['sg_tc'])
    data['ac_oda'].update(data['ac_tc'])
    return data

def CDH_API_muaoda():
    now = datetime.now()
    kt = datetime(now.year,now.month,now.day,7)
    bd = kt - timedelta(hours=25)
    dfname = pd.read_csv('TS_ID/CDH/CDH_MUA_ODA.txt',sep="\s+",header=None)
    data = pd.DataFrame()
    data['time'] = pd.date_range(bd,kt,freq='H')
    for tsid in zip(dfname[0],dfname[1]):
        url = 'https://cdh.vnmha.gov.vn/KiWIS/KiWIS?http://slportal.kttv.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request=getTimeseriesValues&datasource=0&format=html&ts_id={}&from={}&to={}'
        df = pd.read_html(url.format(tsid[0],bd.strftime('%Y-%m-%d'),kt.strftime('%Y-%m-%d')))
        for i in df:
            i = i.iloc[4:,:].dropna()
            i.rename(columns={0:'time',1:tsid[1]},inplace=True)
            i['time'] = pd.to_datetime(i['time'])
            i['time'] = i['time'].dt.strftime('%Y-%m-%d %H:%M:%S')
            i['time'] = pd.to_datetime(i['time'])
            data = data.merge(i,how='left',on='time')
    data.set_index('time',inplace=True)
    data =data.astype(float)
    data.fillna(method='ffill', inplace=True) # thay the nhung gia tri trong bang nan
    data = data.diff()
    tgg = bd + timedelta(hours=1)
    data = data.loc[data.index >= tgg]
    return data
   
def CDH_API_muavrain():
    now = datetime.now()
    kt = datetime(now.year,now.month,now.day,7)
    bd = kt - timedelta(hours=25)
    dfname = pd.read_csv('TS_ID/CDH/CDH_MUA_VRAIN.txt',sep="\s+",header=None)
    data = pd.DataFrame()
    data['time'] = pd.date_range(bd,kt,freq='T')
    for tsid in zip(dfname[0],dfname[1]):
        url = 'https://cdh.vnmha.gov.vn/KiWIS/KiWIS?http://slportal.kttv.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request=getTimeseriesValues&datasource=0&format=html&ts_id={}&from={}&to={}'
        df = pd.read_html(url.format(tsid[0],bd.strftime('%Y-%m-%d'),kt.strftime('%Y-%m-%d')))
        for i in df:
            i = i.iloc[4:,:].dropna()
            i.rename(columns={0:'time',1:tsid[1]},inplace=True)
            i['time'] = pd.to_datetime(i['time'])
            i['time'] = i['time'].dt.strftime('%Y-%m-%d %H:%M:%S')
            i['time'] = pd.to_datetime(i['time'])
            data = data.merge(i,how='left',on='time')
    data.set_index('time',inplace=True)
    muagio = data.rolling(60,min_periods=1).sum()
    muagio = muagio[muagio.index.minute == 0]
    muagio =muagio.astype(float)
    tgg = bd + timedelta(hours=1)
    muagio = muagio.loc[muagio.index >= tgg]
    return muagio
        
        
def tinhdactrungngay():
    pth = tim_file(read_txt('path_tin/DATA_EXCEL.txt'),'.xlsm')
    now = datetime.now()
    now = datetime(now.year,now.month,now.day,7)
    kt = now - timedelta(days=1)
    df = pd.read_excel(pth,sheet_name='H')
    df.rename(columns={'Ngày':'time','Trà Bồng\n(Châu Ổ)':'Châu Ổ'},inplace=True)
    # print(df)
    dt_rang = pd.date_range(start=datetime(2022,1,1,1), periods=len(df['time']), freq="H")
    df['time'] = dt_rang
    data = df.loc[(df['time'] > kt) & (df['time'] <= now )]
    data['An Chỉ']= data['An Chỉ'].interpolate(method='linear')
    # print(data)
        # ghi so loc so lieu
    wb = load_workbook(pth,keep_vba=True)
    ws = wb['hangngay']
    # thuc do max min
    ws['G5'] = data['Trà Khúc'].max()
    ws['G7'] = data['Sông Vệ'].max()
    ws['G8'] = data['Châu Ổ'].max()
    ws['G9'] = data['Trà Câu'].max()

    ws['H5'] = data['Trà Khúc'].min()
    ws['H7'] = data['Sông Vệ'].min()
    ws['H8'] = data['Châu Ổ'].min()
    ws['H9'] = data['Trà Câu'].min()

    # du bao max min
    ws['P5'] = data['Trà Khúc'].max()
    ws['P7'] = data['Sông Vệ'].max()
    ws['P8'] = data['Châu Ổ'].max()
    ws['P9'] = data['Trà Câu'].max()

    ws['Q5'] = data['Trà Khúc'].min()
    ws['Q7'] = data['Sông Vệ'].min()
    ws['Q8'] = data['Châu Ổ'].min()
    ws['Q9'] = data['Trà Câu'].min()

    # tram khong anh huong trieu
    # data = data.loc[data[cotdainhat].dt.hour == 7]
    tn = data.tail(1)
    ws['F4'] = tn['Sơn Giang'].iloc[0]
    ws['F6'] = tn['An Chỉ'].iloc[0]


    #12h ngay hom qua
    h12 = data.loc[data['time'].dt.hour == 12]
    ws['L4'] = h12['Sơn Giang'].iloc[0]
    ws['L5'] = h12['Trà Khúc'].iloc[0]
    ws['L6'] = h12['An Chỉ'].iloc[0]
    ws['L7'] = h12['Sông Vệ'].iloc[0]  
    ws['L8'] = h12['Châu Ổ'].iloc[0]
    ws['L9'] = h12['Trà Câu'].iloc[0]
   
    h18 = data.loc[data['time'].dt.hour == 18]
    ws['M4'] = h18['Sơn Giang'].iloc[0]
    ws['M5'] = h18['Trà Khúc'].iloc[0]
    ws['M6'] = h18['An Chỉ'].iloc[0]
    ws['M7'] = h18['Sông Vệ'].iloc[0]  
    ws['M8'] = h18['Châu Ổ'].iloc[0]
    ws['M9'] = h18['Trà Câu'].iloc[0]

    h0 = data.loc[data['time'].dt.hour == 0]
    ws['N4'] = h0['Sơn Giang'].iloc[0]
    ws['N5'] = h0['Trà Khúc'].iloc[0]
    ws['N6'] = h0['An Chỉ'].iloc[0]
    ws['N7'] = h0['Sông Vệ'].iloc[0]  
    ws['N8'] = h0['Châu Ổ'].iloc[0]
    ws['N9'] = h0['Trà Câu'].iloc[0]

    h7 = data.loc[data['time'].dt.hour == 7]
    ws['O4'] = h7['Sơn Giang'].iloc[0]
    ws['O5'] = h7['Trà Khúc'].iloc[0]
    ws['O6'] = h7['An Chỉ'].iloc[0]
    ws['O7'] = h7['Sông Vệ'].iloc[0]  
    ws['O8'] = h7['Châu Ổ'].iloc[0]
    ws['O9'] = h7['Trà Câu'].iloc[0]
    wb.save(pth)
    time.sleep(3)
    excel = client.Dispatch("Excel.Application")
    excel.Visible = True
    book = excel.Workbooks.Open(pth)
    book.Worksheets('H').Select()
  
def hochua():
    now = datetime.now()
    kt = datetime(now.year,now.month,now.day,7)
    bd = kt - timedelta(days=1)
    dfname = pd.read_csv('TS_ID/CDH/CDH_H_hochua.txt',sep="\s+",header=None)
    data = pd.DataFrame()
    data['time'] = pd.date_range(bd,kt,freq='H')
    for tsid in zip(dfname[0],dfname[1]):
        url = 'https://cdh.vnmha.gov.vn/KiWIS/KiWIS?http://slportal.kttv.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request=getTimeseriesValues&datasource=0&format=html&ts_id={}&from={}&to={}'
        df = pd.read_html(url.format(tsid[0],bd.strftime('%Y-%m-%d'),kt.strftime('%Y-%m-%d')))
        for i in df:
            i = i.iloc[4:,:].dropna()
            i.rename(columns={0:'time',1:tsid[1]},inplace=True)
            i['time'] = pd.to_datetime(i['time'])
            i['time'] = i['time'].dt.strftime('%Y-%m-%d %H:%M:%S')
            i['time'] = pd.to_datetime(i['time'])
            data = data.merge(i,how='left',on='time')
    data.set_index('time',inplace=True)
    data =data.astype(float)
    data.fillna(method='ffill', inplace=True)
    pth = tim_file(read_txt('path_tin/DATA_EXCEL.txt'),'.xlsm')

    wb = load_workbook(pth,keep_vba=True)
    ws = wb['tuan5ngay']
    ws['M12'] = data['Hdr'].iloc[-2]
    ws['M13'] = data['Hnt'].tail(1).values[0]
    wb.save(pth)
    messagebox.showinfo('Thông báo','Đã xong!')

def get_TTB_API():
    data_H = TTB_API_mucnuoc()
    data_mua =TTB_API_mua()
    id = vitridat() # tim vi tri dat
    pth = tim_file(read_txt('path_tin/DATA_EXCEL.txt'),'.xlsm')
    # pth = 'Data.xlsm'
    with pd.ExcelWriter(pth,mode='a',engine_kwargs={'keep_vba': True},engine='openpyxl',if_sheet_exists='overlay') as writer:   # ghi vao file co san
        data_H.to_excel(writer, sheet_name='H',startrow=id -1, startcol=1, header=False, index=False)
        data_mua.to_excel(writer, sheet_name='Mua',startrow=id -1, startcol=0, header=False, index=True)     
    tinhdactrungngay()
    messagebox.showinfo('Thông báo',"Đã load xong!")
    
def get_CDH_API():
    data_H = CDH_API_mucnuoc()
    data_mua_oda = CDH_API_muaoda()
    data_mua_vrain = CDH_API_muavrain()
    id = vitridat() # tim vi tri dat
    pth = tim_file(read_txt('path_tin/DATA_EXCEL.txt'),'.xlsm')
    with pd.ExcelWriter(pth,mode='a',engine_kwargs={'keep_vba': True},engine='openpyxl',if_sheet_exists='overlay') as writer:   # ghi vao file co san
        data_H.to_excel(writer, sheet_name='H',startrow=id -1, startcol=1, header=False, index=False)
        data_mua_oda.to_excel(writer, sheet_name='Mua',startrow=id -1, startcol=0, header=False, index=True)
        data_mua_vrain.to_excel(writer, sheet_name='Mua',startrow=id -1, startcol=16, header=False, index=False)
    tinhdactrungngay()
    messagebox.showinfo('Thông báo',"Đã load xong!")


    
